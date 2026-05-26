import sys
import os
import pandas as pd
import subprocess
from copy import copy
from PyQt6.QtWidgets import (
    QApplication,
    QWidget,
    QPushButton,
    QLabel,
    QFileDialog,
    QMessageBox,
    QVBoxLayout,
    QHBoxLayout,
    QLineEdit,
    QFrame,
    QCheckBox
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import (QIcon, QPixmap)
from openpyxl import load_workbook
from openpyxl.styles import (
    Font,
    Border,
    Side,
    PatternFill,
    Alignment
)
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import (
    AnchorMarker,
    OneCellAnchor
)

# =========================================================
# CONFIG
# =========================================================

OPERACOES_VALIDAS = [
    "FOLHA PAGAMENTO - BOLSISTAS", "FOLHA PAGAMENTO - BOLSISTAS ",
    "FOLHA PAGAMENTO - ALUNOS BOLSISTAS", "FOLHA PAGAMENTO - ALUNOS BOLSISTAS "
]

TEMPLATE_FILE = "template.xlsx"

# =========================================================
# AUXILIARES
# =========================================================

def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS

    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

def copiar_estilo(origem, destino):

    if origem.has_style:

        destino.font = copy(origem.font)
        destino.fill = copy(origem.fill)
        destino.border = copy(origem.border)
        destino.alignment = copy(origem.alignment)
        destino.number_format = copy(origem.number_format)
        destino.protection = copy(origem.protection)

def copiar_linha(ws, linha_modelo, linha_destino):

    for col in range(1, ws.max_column + 1):

        origem = ws.cell(linha_modelo, col)
        destino = ws.cell(linha_destino, col)

        copiar_estilo(origem, destino)

    ws.row_dimensions[linha_destino].height = (
        ws.row_dimensions[linha_modelo].height
    )

def detectar_header(df_raw):

    for row in range(min(10, len(df_raw))):

        for col in range(df_raw.shape[1]):

            valor = str(
                df_raw.iat[row, col]
            ).strip()

            if valor == "Operação":
                return row

    raise Exception(
        "Não foi possível localizar a linha de cabeçalho."
    )

def formatar_mes_pagto(data):

    if pd.isna(data):
        return ""

    meses = {
        1: "jan",
        2: "fev",
        3: "mar",
        4: "abr",
        5: "mai",
        6: "jun",
        7: "jul",
        8: "ago",
        9: "set",
        10: "out",
        11: "nov",
        12: "dez"
    }

    mes = meses[data.month]

    return f"{mes}/{str(data.year)[-2:]}"

def cm_para_px(cm):
    return int((cm / 2.54) * 96)

# =========================================================
# PROCESSAMENTO
# =========================================================

def gerar_planilha(arquivo_entrada, titulo_pc, caminho_saida):

    # =====================================================
    # LEITURA
    # =====================================================

    df_raw = pd.read_excel(
        arquivo_entrada,
        header=None
    )

    linha_header = detectar_header(df_raw)

    df = pd.read_excel(
        arquivo_entrada,
        header=linha_header
    )

    # =====================================================
    # FILTRO
    # =====================================================

    df = df[
        df["Operação"].isin(
            OPERACOES_VALIDAS
        )
    ].copy()

    # =====================================================
    # DATAS
    # =====================================================

    df["Emissão"] = pd.to_datetime(
        df["Emissão"],
        dayfirst=True,
        errors="coerce"
    )

    df["Vencimento/Mov."] = pd.to_datetime(
        df["Vencimento/Mov."],
        dayfirst=True,
        errors="coerce"
    )

    # =====================================================
    # TEMPLATE
    # =====================================================

    wb = load_workbook(
        resource_path(TEMPLATE_FILE)
    )

    ws_rosto = wb["Folha de Rosto"]
    ws_det = wb["Detalhamento"]

    EMU_PER_PIXEL = 9525

    header = resource_path("header.png")

    if os.path.exists(header):
        # FOLHA DE ROSTO
        img_rosto = Image(header)
        img_rosto.height = cm_para_px(6)
        img_rosto.width = cm_para_px(20.82)

        marker_rosto = AnchorMarker(
            col=0,          # coluna A = 0
            colOff=50 * EMU_PER_PIXEL,   # offset horizontal
            row=0,          # linha 1 = 0
            rowOff=30 * EMU_PER_PIXEL    # offset vertical
        )
        img_rosto.anchor = OneCellAnchor(
            _from=marker_rosto,
            ext=None
        )
        ws_rosto.add_image(img_rosto, "A1")

        # DETALHAMENTO
        img_det = Image(header)
        img_det.height = cm_para_px(4.42)
        img_det.width = cm_para_px(16.83)

        marker_det = AnchorMarker(
            col=0,          # coluna A = 0
            colOff=50 * EMU_PER_PIXEL,   # offset horizontal
            row=0,          # linha 1 = 0
            rowOff=30 * EMU_PER_PIXEL    # offset vertical
        )
        img_det.anchor = OneCellAnchor(
            _from=marker_det,
            ext=None
        )
        ws_det.add_image(img_det, "A1")

    else:
        QMessageBox.warning(window,
            "Imagem não encontrada",
            "O cabeçalho da folha de rosto não foi localizado."
        )

    # =====================================================
    # CAMPOS FIXOS
    # =====================================================

    ws_rosto["E18"] = titulo_pc
    ws_det["C17"] = titulo_pc

    # =====================================================
    # FOLHA DE ROSTO
    # =====================================================

    linha_inicio_rosto = 25
    linha_modelo_rosto = 25

    linha_atual = linha_inicio_rosto

    total_rosto = 0

    for _, row in df.iterrows():

        if linha_atual > linha_modelo_rosto:
            copiar_linha(
                ws_rosto,
                linha_modelo_rosto,
                linha_atual
            )

        nf = row.get("Nº NF")

        if pd.isna(nf):
            nf = "-"

        emissao = row.get("Emissão")
        venc = row.get("Vencimento/Mov.")
        valor = abs(row.get("Valor Total"))

        ws_rosto[f"C{linha_atual}"] = nf

        ws_rosto[f"D{linha_atual}"] = emissao
        ws_rosto[f"D{linha_atual}"].number_format = "dd/mm/yyyy"

        ws_rosto[f"E{linha_atual}"] = formatar_mes_pagto(venc)

        ws_rosto[f"F{linha_atual}"] = titulo_pc

        ws_rosto[f"G{linha_atual}"] = venc
        ws_rosto[f"G{linha_atual}"].number_format = "dd/mm/yyyy"

        ws_rosto[f"H{linha_atual}"] = valor
        ws_rosto[f"H{linha_atual}"].number_format = (
            'R$ #,##0.00'
        )

        total_rosto += (
            valor if pd.notna(valor)
            else 0
        )

        linha_atual += 1

    # =====================================================
    # TOTAL ROSTO
    # =====================================================

    total_row = linha_atual

    ws_rosto[f"H{total_row}"] = (
        f"=SUM(H25:H{total_row - 1})"
    )

    cell = ws_rosto[f"H{total_row}"]

    cell.font = Font(
        name="Calibri",
        size=11,
        bold=True
    )

    cell.fill = PatternFill(
        fill_type="solid",
        start_color="D9D9D9",
        end_color="D9D9D9"
    )

    thick = Side(style="thick")

    cell.border = Border(
        left=thick,
        right=thick,
        top=thick,
        bottom=thick
    )

    cell.alignment = Alignment(
        horizontal="right",
        vertical="center",
        wrap_text=True
    )

    cell.number_format = 'R$ #,##0.00'

    # =====================================================
    # DETALHAMENTO
    # =====================================================

    linha_inicio_det = 49
    linha_modelo_det = 49

    linha_atual_det = linha_inicio_det

    for _, row in df.iterrows():

        if linha_atual_det > linha_modelo_det:

            copiar_linha(
                ws_det,
                linha_modelo_det,
                linha_atual_det
            )

        ws_det[f"B{linha_atual_det}"] = (
            row.get("Razão Social")
        )

        ws_det[f"C{linha_atual_det}"] = (
            row.get("CPF/CNPJ")
        )

        ws_det[f"D{linha_atual_det}"] = abs(
            row.get("Valor Total")
        )

        ws_det[f"D{linha_atual_det}"].number_format = (
            'R$ #,##0.00'
        )

        linha_atual_det += 1

    # =====================================================
    # TOTAL DETALHAMENTO
    # =====================================================

    total_det_row = linha_atual_det

    copiar_estilo(
        ws_det[f"D{total_det_row - 1}"],
        ws_det[f"D{total_det_row}"]
    )

    ws_det[f"D{total_det_row}"] = (
        f"=SUM(D49:D{total_det_row - 1})"
    )

    ws_det[f"D{total_det_row}"].number_format = (
        'R$ #,##0.00'
    )

    # =====================================================
    # SAVE
    # =====================================================

    wb.save(caminho_saida)

# =========================================================
# UI
# =========================================================

class ToggleSwitch(QCheckBox):

    def __init__(self, text="", parent=None):

        super().__init__(text, parent)

        self.setFixedHeight(28)

        self.setCursor(
            Qt.CursorShape.PointingHandCursor
        )

    def paintEvent(self, event):

        from PyQt6.QtGui import (
            QPainter,
            QColor
        )

        from PyQt6.QtCore import QRectF

        painter = QPainter(self)

        painter.setRenderHint(
            QPainter.RenderHint.Antialiasing
        )

        # ============================================
        # TRACK
        # ============================================

        track_rect = QRectF(
            0,
            4,
            44,
            20
        )

        if self.isChecked():

            painter.setBrush(
                QColor("#f9b02e")
            )

        else:

            painter.setBrush(
                QColor("#3a3a3a")
            )

        painter.setPen(
            Qt.PenStyle.NoPen
        )

        painter.drawRoundedRect(
            track_rect,
            10,
            10
        )

        # ============================================
        # KNOB
        # ============================================

        knob_x = (
            24
            if self.isChecked()
            else 2
        )

        knob_rect = QRectF(
            knob_x,
            6,
            16,
            16
        )

        painter.setBrush(
            QColor("white")
        )

        painter.drawEllipse(
            knob_rect
        )

        # ============================================
        # TEXTO
        # ============================================

        painter.setPen(
            QColor("#dddddd")
        )

        painter.drawText(
            54,
            19,
            self.text()
        )

class App(QWidget):

    def __init__(self):

        super().__init__()

        self.input_file = None

        self.init_ui()

    # =====================================================
    # UI
    # =====================================================

    def init_ui(self):

        self.setWindowTitle("SmartPC RH")

        self.setWindowIcon(
            QIcon(resource_path("smartpc_rh.ico"))
        )

        self.resize(960, 540)

        self.setMinimumSize(960, 540)

        # =================================================
        # STYLE
        # =================================================

        self.setStyleSheet("""
        QWidget {
            background-color: #1e1e1e;
            font-family: "Segoe UI";
            color: white;
        }

        QLineEdit {
            background-color: #2b2b2b;
            color: white;
            border: 1px solid #3a3a3a;
            border-radius: 8px;
            padding: 8px;
            font-size: 13px;
        }

        QLineEdit:focus {
            border: 1px solid #f9b02e;
        }
        """)

        # =================================================
        # BACKGROUND IMAGE
        # =================================================

        self.bg_label = QLabel(self)

        bg_pixmap = QPixmap(
            resource_path("bg_hbr.png")
        )

        self.bg_label.setPixmap(bg_pixmap)

        self.bg_label.setScaledContents(True)

        # =================================================
        # BUTTON STYLE
        # =================================================

        self.button_style = """
        QPushButton {
            background-color: #f9b02e;
            color: black;
            border: none;
            border-radius: 8px;
            padding: 10px;
            font-weight: bold;
            font-size: 14px;
        }

        QPushButton:hover {
            background-color: #ffd166;
        }

        QPushButton:pressed {
            background-color: #e69500;
        }

        QPushButton:disabled {
            background-color: #666666;
            color: #aaaaaa;
        }
        """

        # =================================================
        # MAIN LAYOUT
        # =================================================

        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(
            30,
            30,
            30,
            30
        )
        main_layout.setSpacing(12)
        main_layout.addStretch()

        # =================================================
        # CARD
        # =================================================

        card = QFrame()

        card.setObjectName("card")

        card.setStyleSheet("""
        #card {
            background-color: rgba(30,30,30,220);
            border-radius: 18px;
            border: 1px solid rgba(255,255,255,25);
        }
        """)

        card.setMaximumWidth(620)

        main_layout.addWidget(
            card,
            alignment=Qt.AlignmentFlag.AlignCenter
        )

        main_layout.addStretch()

        # =================================================
        # CARD LAYOUT
        # =================================================

        card_layout = QVBoxLayout(card)

        card_layout.setContentsMargins(
            24,
            24,
            24,
            24
        )

        card_layout.setSpacing(14)

        # =================================================
        # TITLE
        # =================================================

        titulo = QLabel(
            "SmartPC RH"
        )

        titulo.setAlignment(
            Qt.AlignmentFlag.AlignCenter
        )

        titulo.setStyleSheet("""
        font-family: "Bahnschrift Condensed";
        font-size: 38px;
        font-weight: bold;
        color: white;
        background: transparent;
        """)

        card_layout.addWidget(titulo)

        # =================================================
        # LABEL FILE
        # =================================================

        self.label_file = QLabel("-")

        # =================================================
        # BLOCO ARQUIVO
        # =================================================

        card_layout.addWidget(
            self.criar_bloco_arquivo(
                "Relatório",
                self.selecionar_arquivo,
                self.remover_arquivo,
                self.label_file
            )
        )

        # =================================================
        # INPUT CARD
        # =================================================

        input_card = QFrame()

        input_card.setStyleSheet("""
        QFrame {
            background-color: rgba(255,255,255,8);
            border-radius: 12px;
        }
        """)

        input_layout = QVBoxLayout(input_card)

        input_layout.setContentsMargins(
            14,
            14,
            14,
            14
        )

        input_layout.setSpacing(10)

        pc_title = QLabel(
            "Título PC"
        )

        pc_title.setStyleSheet("""
        color: white;
        font-size: 13px;
        font-weight: bold;
        background: transparent;
        """)

        self.input_pc = QLineEdit()

        self.input_pc.setPlaceholderText(
            "Digite o título da PC:"
        )

        self.input_pc.setFixedHeight(38)

        input_layout.addWidget(pc_title)
        input_layout.addWidget(self.input_pc)

        card_layout.addWidget(input_card)

        # =================================================
        # TOGGLE
        # =================================================

        self.checkbox_abrir = ToggleSwitch(
            "Abrir planilha quando estiver pronta"
        )

        card_layout.addWidget(
            self.checkbox_abrir
        )

        # =================================================
        # BOTÃO GERAR
        # =================================================

        self.btn_exec = QPushButton(
            "Gerar Planilha"
        )

        self.btn_exec.setEnabled(False)

        self.btn_exec.setFixedSize(
            170,
            42
        )

        self.btn_exec.setStyleSheet(
            self.button_style
        )

        self.btn_exec.clicked.connect(
            self.executar
        )

        gerar_layout = QHBoxLayout()

        gerar_layout.addStretch()
        gerar_layout.addWidget(self.btn_exec)
        gerar_layout.addStretch()

        card_layout.addLayout(
            gerar_layout
        )

        # =================================================
        # FOOTER
        # =================================================

        github_label = QLabel()
        github_label.setAlignment(Qt.AlignmentFlag.AlignCenter)

        github_label.setText(
            '<a href="https://github.com/imbaTIMvel/smartpc_rh">'
            'SmartPC RH v0.1.0 - GitHub'
            '</a>'
        )

        github_label.setOpenExternalLinks(True)

        github_label.setStyleSheet("""
        QLabel {
            background-color: transparent;
            color: rgba(255,255,255,120);
            font-size: 11px;
        }

        QLabel:hover {
            color: #f9b02e;
        }
        """)

        footer = QLabel(
            "Desenvolvido por: Diretoria Administrativa Financeira - DAF"
        )

        footer.setAlignment(Qt.AlignmentFlag.AlignCenter)

        footer.setStyleSheet("""
        QLabel {
            background-color: transparent;
            color: rgba(255,255,255,120);
            font-size: 10px;
            padding-bottom: 4px;
        }
        """)

        # =================================================
        # WINDOW LAYOUT
        # =================================================

        window_layout = QVBoxLayout(self)

        window_layout.setContentsMargins(
            0,
            0,
            0,
            0
        )

        window_layout.addLayout(main_layout)
        window_layout.addWidget(github_label, alignment=Qt.AlignmentFlag.AlignBottom)
        window_layout.addWidget(footer, alignment=Qt.AlignmentFlag.AlignBottom)

    def resizeEvent(self, event):

        self.bg_label.resize(self.size())

        super().resizeEvent(event)

    # =====================================================
    # BLOCO
    # =====================================================

    def criar_bloco_arquivo(
        self,
        titulo,
        callback_select,
        callback_remove,
        label_widget
    ):

        bloco = QFrame()

        bloco.setStyleSheet("""
        QFrame {
            background-color: rgba(255,255,255,8);
            border-radius: 12px;
        }
        """)

        bloco_layout = QVBoxLayout(bloco)

        bloco_layout.setContentsMargins(
            14,
            14,
            14,
            14
        )

        bloco_layout.setSpacing(10)

        titulo_label = QLabel(titulo)

        titulo_label.setStyleSheet("""
        color: white;
        font-size: 13px;
        font-weight: bold;
        background: transparent;
        """)

        bloco_layout.addWidget(titulo_label)

        botoes_layout = QHBoxLayout()

        btn_select = QPushButton("Selecionar")

        btn_select.setFixedSize(120, 34)

        btn_select.setStyleSheet(
            self.button_style
        )

        btn_select.clicked.connect(
            callback_select
        )

        btn_remove = QPushButton("Remover")

        btn_remove.setFixedSize(120, 34)

        btn_remove.setStyleSheet(
            self.button_style
        )

        btn_remove.clicked.connect(
            callback_remove
        )

        botoes_layout.addWidget(btn_select)
        botoes_layout.addWidget(btn_remove)
        botoes_layout.addStretch()

        bloco_layout.addLayout(
            botoes_layout
        )

        label_widget.setWordWrap(True)

        label_widget.setStyleSheet("""
        background-color: #2b2b2b;
        border: 1px solid #3a3a3a;
        border-radius: 8px;
        padding: 8px;
        color: #cccccc;
        font-size: 11px;
        """)

        label_widget.setMinimumHeight(20)

        bloco_layout.addWidget(
            label_widget
        )

        return bloco

    # =====================================================
    # SELECIONAR
    # =====================================================
    
    def selecionar_arquivo(self):
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Selecionar arquivo",
            "",
            "Excel (*.xlsx)"
            )
        
        if path:
            self.input_file = path
            self.label_file.setText(
                os.path.basename(path)
                )
            self.btn_exec.setEnabled(True)

    # =====================================================
    # REMOVER
    # =====================================================

    def remover_arquivo(self):

        self.input_file = None

        self.label_file.setText("-")

        self.btn_exec.setEnabled(False)

    # =====================================================
    # EXECUTAR
    # =====================================================
    
    def executar(self):
        try:
            if not self.input_pc.text().strip():
                QMessageBox.warning(
                    self,
                    "Erro",
                    "Digite a prestação de contas."
                    )
                return
            
            QMessageBox.information(
                self,
                "Sucesso",
                "Planilha criada com sucesso!"
                )

            save_path, _ = QFileDialog.getSaveFileName(
                self,
                "Salvar arquivo",
                "",
                "Excel (*.xlsx)"
                )
            
            if not save_path:
                return
            
            gerar_planilha(
                self.input_file,
                self.input_pc.text(),
                save_path
                )
            
            if self.checkbox_abrir.isChecked():
                try:
                    os.startfile(save_path)
                except:
                    subprocess.call(
                        ["open",
                         save_path]
                         )
        
        except Exception as e:
            QMessageBox.critical( self, "Erro", str(e) )

# =========================================================
# MAIN
# =========================================================

app = QApplication(sys.argv)

window = App()

window.show()

sys.exit(app.exec())